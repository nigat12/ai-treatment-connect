import pandas as pd
import random
import calendar
from datetime import date, timedelta
import openpyxl # Required by pandas for xlsx writing
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
import platform # To handle date formatting differences

# --- Configuration ---

# Agent data (Division, Name) - Add more as needed
AGENTS_DATA = [
    ("East", "Demelash Gebru Wolde"),
    ("East", "Menjo Trading PLC"),
    ("East", "Esrael Teshome"),
    ("East", "Miliyon Hailu Biza"),
    ("East", "Naol Gebreyohannis Gonfa"),
    ("East", "Tadese Bogale"),
    ("East", "Tatek Legesse Belachew"),
    ("East", "Shiferaw Bekele Dema"), # Agent with potentially zero targets
    ("East", "Mulu Tilaye Ayele"),
]

# Brands (used for columns in summary and for subsequent tabs)
BRANDS = ["Habesha", "Kidame", "Negus", "Feta"]
ALL_BRANDS_WITH_KOSTARA = BRANDS + ["Kostara"] # For summary tab header order

# Mapping for tab names and last column header in daily tabs
BRAND_TAB_DETAILS = {
    "All": "ALL",
    "Habesha": "Habesha Beer",
    "Kidame": "Kidame Beer",
    "Negus": "NEGUS",
    "Feta": "Feta Beer"
}
# Add 'All' to the list of tabs to generate
ALL_DAILY_TABS = ["All"] + BRANDS

NUM_TRUCKS_PER_AGENT = 2

# --- NEW: Random DAILY target ranges ---
MIN_DAILY_TARGET_PER_BRAND_TRUCK = 5   # Min target for a single day/brand/truck
MAX_DAILY_TARGET_PER_BRAND_TRUCK = 250 # Max target for a single day/brand/truck
# Probability that a specific brand gets *zero* daily targets for an agent/truck
# (Applied *before* generating random numbers for that brand/agent/truck combo)
ZERO_BRAND_PROBABILITY_DAILY = 0.10 # 10% chance a brand has no daily targets for a truck

# --- Helper Functions ---

def get_month_year():
    """Gets and validates month and year input from the user."""
    while True:
        try:
            month = int(input("Enter the month number (1-12): "))
            if 1 <= month <= 12:
                break
            else:
                print("Invalid month. Please enter a number between 1 and 12.")
        except ValueError:
            print("Invalid input. Please enter a number.")
    while True:
        try:
            year = int(input("Enter the year (e.g., 2025): "))
            if 1900 < year < 3000:
                break
            else:
                print("Invalid year. Please enter a reasonable year.")
        except ValueError:
            print("Invalid input. Please enter a number.")
    return month, year

def get_dates_in_month(year, month):
    """Returns a list of all date objects in the given month and year."""
    num_days = calendar.monthrange(year, month)[1]
    start_date = date(year, month, 1)
    return [start_date + timedelta(days=i) for i in range(num_days)]

def generate_truck_names(agent_name, num_trucks):
    """Generates unique-ish truck names for an agent."""
    # Simple generation, ensure consistency if called multiple times for same agent
    random.seed(hash(agent_name)) # Seed based on agent name for reproducibility per agent
    base_hash = abs(hash(agent_name)) % 10000
    names = []
    for i in range(num_trucks):
        random_part = random.randint(10000, 99999)
        truck_id = f"Truck {base_hash % 10}-{random_part}"
        names.append(truck_id)
        base_hash += i # Increment consistently
    random.seed() # Reset seed
    return names

# --- Main Logic ---

def generate_sales_target_excel(month, year):
    """Generates the Excel file with sales targets using a bottom-up approach."""

    month_name = calendar.month_abbr[month]
    filename = f"{month_name}.{year}.Target.xlsx"
    print(f"Generating target file: {filename}")

    # Determine date format based on OS ('%-m/%-d/%Y' for Linux/Mac, '%#m/%#d/%Y' for Windows)
    date_format = '%#m/%#d/%Y' if platform.system() == "Windows" else '%-m/%-d/%Y'


    all_dates = get_dates_in_month(year, month)

    # == 1. Generate Base Daily Targets ==
    # Structure: base_daily_targets[agent_name][truck_name][brand_name][date_obj] = numeric_target
    print("Generating base daily targets...")
    base_daily_targets = {}

    agent_truck_map = { # Store truck names per agent
        agent_name: generate_truck_names(agent_name, NUM_TRUCKS_PER_AGENT)
        for _, agent_name in AGENTS_DATA
    }

    for division, agent_name in AGENTS_DATA:
        base_daily_targets[agent_name] = {}
        agent_trucks = agent_truck_map[agent_name]

        for truck in agent_trucks:
            base_daily_targets[agent_name][truck] = {}
            for brand in BRANDS:
                base_daily_targets[agent_name][truck][brand] = {}

                # Decide if this brand gets *any* targets for this truck/agent
                if agent_name == "Shiferaw Bekele Dema" or random.random() < ZERO_BRAND_PROBABILITY_DAILY:
                     # Assign 0 for all working days for this specific brand/truck/agent
                     for dt in all_dates:
                         if dt.weekday() < 6: # Working day
                             base_daily_targets[agent_name][truck][brand][dt] = 0
                     continue # Skip random generation for this brand/truck

                # Generate random target for each working day
                for dt in all_dates:
                    if dt.weekday() < 6: # If it's a working day (Mon-Sat)
                        target = random.randint(MIN_DAILY_TARGET_PER_BRAND_TRUCK, MAX_DAILY_TARGET_PER_BRAND_TRUCK)
                        base_daily_targets[agent_name][truck][brand][dt] = target
                    # Sundays will be handled later (implicitly 0 here, displayed as '-')


    # == 2 & 3. Populate Daily Tab Data (Individual Brands and 'All') ==
    print("Populating daily tab data...")
    # daily_tab_data[tab_name] = list_of_rows (where each row is a dictionary)
    daily_tab_data = {tab_name: [] for tab_name in ALL_DAILY_TABS}

    for division, agent_name in AGENTS_DATA:
        agent_trucks = agent_truck_map[agent_name]
        for truck in agent_trucks:
            for dt in all_dates:
                date_str = dt.strftime(date_format)
                is_sunday = (dt.weekday() == 6)
                daily_all_brands_sum = 0
                has_any_numeric_target = False

                # Populate individual brand tabs first
                for brand in BRANDS:
                    brand_tab_name = brand
                    brand_col_header = BRAND_TAB_DETAILS[brand]

                    numeric_target = 0 # Default to 0
                    if not is_sunday:
                         # Retrieve the pre-generated base target
                         numeric_target = base_daily_targets.get(agent_name, {}).get(truck, {}).get(brand, {}).get(dt, 0)

                    # Determine display value: '-' for Sunday or if target is 0, else the number
                    display_value = '-' if (is_sunday or numeric_target == 0) else numeric_target

                    # Store row for the individual brand tab
                    row_data_brand = {
                        "Date": date_str,
                        "Division": division,
                        "Sales Unit": agent_name,
                        "Truck": truck,
                        brand_col_header: display_value
                    }
                    daily_tab_data[brand_tab_name].append(row_data_brand)

                    # Add to the sum for the 'All' tab calculation (only if not Sunday)
                    if not is_sunday and numeric_target > 0:
                        daily_all_brands_sum += numeric_target
                        has_any_numeric_target = True

                # Populate 'All' tab row for this date/agent/truck
                all_tab_col_header = BRAND_TAB_DETAILS["All"]
                # Display '-' on Sunday OR if the sum of targets is 0 for that day
                all_display_value = '-' if (is_sunday or not has_any_numeric_target) else daily_all_brands_sum

                row_data_all = {
                    "Date": date_str,
                    "Division": division,
                    "Sales Unit": agent_name,
                    "Truck": truck,
                    all_tab_col_header: all_display_value
                }
                daily_tab_data["All"].append(row_data_all)

    # Convert lists of rows into DataFrames
    daily_data_frames = {
        tab_name: pd.DataFrame(daily_tab_data[tab_name])
        for tab_name in ALL_DAILY_TABS
    }

    # == 4. Calculate and Populate Monthly Summary Tab ==
    print("Calculating monthly summary data...")
    summary_data = []
    # Structure: monthly_summary_totals[agent_name][brand_name] = total_monthly_target
    monthly_summary_totals = {}

    for division, agent_name in AGENTS_DATA:
        agent_monthly_data = {"Division": division, "Agent Name": agent_name}
        agent_total_all_brands = 0
        monthly_summary_totals[agent_name] = {}

        for brand in BRANDS:
            brand_monthly_total = 0
            agent_trucks = agent_truck_map[agent_name]
            for truck in agent_trucks:
                for dt in all_dates:
                    if dt.weekday() < 6: # Only sum working days
                         # Sum the base daily targets generated earlier
                         brand_monthly_total += base_daily_targets.get(agent_name, {}).get(truck, {}).get(brand, {}).get(dt, 0)

            # Store the numeric total
            monthly_summary_totals[agent_name][brand] = brand_monthly_total
            # Set display value for the summary table ('-' if 0)
            agent_monthly_data[brand] = brand_monthly_total if brand_monthly_total > 0 else '-'
            agent_total_all_brands += brand_monthly_total

        # Kostara is always '-'
        agent_monthly_data["Kostara"] = '-'
        # Calculate Total, display '-' if sum is 0
        agent_monthly_data["Total, All |SKU"] = agent_total_all_brands if agent_total_all_brands > 0 else '-'
        monthly_summary_totals[agent_name]['Total'] = agent_total_all_brands # Store numeric total

        summary_data.append(agent_monthly_data)

    # Create Summary DataFrame
    summary_df = pd.DataFrame(summary_data)

    # Reorder columns
    summary_cols_ordered = ["Division", "Agent Name"] + ALL_BRANDS_WITH_KOSTARA + ["Total, All |SKU"]
    summary_df = summary_df[summary_cols_ordered]

    # Calculate Subtotals using the stored numeric monthly totals
    subtotals = {"Division": "Sub Total", "Agent Name": ""}
    grand_total_all = 0
    for brand in BRANDS:
        brand_subtotal = sum(monthly_summary_totals[agent_name].get(brand, 0) for agent_name in monthly_summary_totals)
        subtotals[brand] = brand_subtotal # Keep numeric for now
        grand_total_all += brand_subtotal

    subtotals["Kostara"] = '-'
    subtotals["Total, All |SKU"] = grand_total_all

    # Create subtotal DataFrame (convert 0s to '-' for display just before writing)
    subtotal_df = pd.DataFrame([subtotals])
    summary_df = pd.concat([summary_df, subtotal_df], ignore_index=True)


    # == 5. Write to Excel File ==
    print(f"Writing data to Excel file: {filename}")
    summary_sheet_name = f"{month_name}.{year}.Target"

    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # --- Write Summary Tab ---
        workbook = writer.book
        if 'Sheet' in workbook.sheetnames: del workbook['Sheet'] # Remove default sheet
        if summary_sheet_name in workbook.sheetnames: del workbook[summary_sheet_name]
        summary_sheet = workbook.create_sheet(title=summary_sheet_name, index=0)

        # Title
        title_cell = summary_sheet['A1']
        title_cell.value = f"{month_name.upper()} Target - {year}"
        title_cell.font = Font(bold=True, size=14); title_cell.alignment = Alignment(horizontal='center', vertical='center')
        summary_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(summary_cols_ordered))
        summary_sheet.row_dimensions[1].height = 20

        # Write DataFrame (Headers on row 3)
        summary_df.to_excel(writer, sheet_name=summary_sheet_name, startrow=2, index=False, header=True)
        summary_sheet = writer.sheets[summary_sheet_name] # Get sheet object again

        # Format Summary Headers (Row 3)
        header_font = Font(bold=True); header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"); header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        summary_sheet.row_dimensions[3].height = 30
        for col_num in range(1, len(summary_cols_ordered) + 1):
            cell = summary_sheet.cell(row=3, column=col_num); cell.font = header_font; cell.alignment = header_alignment; cell.fill = header_fill

        # Format Summary Data and Subtotal Area
        number_cols_indices = [i + 1 for i, col_name in enumerate(summary_cols_ordered) if col_name in BRANDS or col_name == "Total, All |SKU"]
        kostara_col_index = summary_cols_ordered.index("Kostara") + 1
        subtotal_row_idx = len(summary_df) + 2 # Row index in Excel (1-based)
        subtotal_font = Font(bold=True); subtotal_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"); center_alignment = Alignment(horizontal='center')

        for row_idx in range(4, subtotal_row_idx + 1): # Loop from first data row through subtotal row
            is_subtotal = (row_idx == subtotal_row_idx)
            for col_idx in range(1, len(summary_cols_ordered) + 1):
                cell = summary_sheet.cell(row=row_idx, column=col_idx)

                if is_subtotal:
                    cell.font = subtotal_font; cell.fill = subtotal_fill; cell.alignment = center_alignment

                # Apply number format or handle '-'
                if col_idx in number_cols_indices:
                    # Check if the value *should* be numeric (even if currently 0 or '-')
                    if isinstance(cell.value, (int, float)):
                        if cell.value == 0:
                             cell.value = '-'
                             cell.number_format = '@' # Text format for '-'
                             cell.alignment = center_alignment # Center the hyphen
                        else:
                             cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
                             # Don't center numbers unless it's subtotal
                             if not is_subtotal:
                                cell.alignment = Alignment(horizontal='right') # Right-align numbers
                    elif cell.value == '-': # If it's already '-', center it
                        cell.alignment = center_alignment
                elif col_idx == kostara_col_index or cell.value == '-': # Center Kostara or any other '-'
                     cell.alignment = center_alignment


        # --- Write Daily Tabs ---
        for tab_name in ALL_DAILY_TABS: # Ensure correct order if needed, but names are distinct
             df = daily_data_frames.get(tab_name)
             if df is not None and not df.empty:
                df.to_excel(writer, sheet_name=tab_name, index=False)
                daily_sheet = writer.sheets[tab_name]
                header_font = Font(bold=True); header_alignment = Alignment(horizontal='center', vertical='center')

                # Format headers
                for col_num in range(1, len(df.columns) + 1):
                    cell = daily_sheet.cell(row=1, column=col_num); cell.font = header_font; cell.alignment = header_alignment

                # Format last column (target column)
                last_col_idx = len(df.columns)
                center_alignment = Alignment(horizontal='center')
                for row_idx in range(2, len(df) + 2): # Data starts row 2
                    cell = daily_sheet.cell(row=row_idx, column=last_col_idx)
                    cell.alignment = center_alignment # Center all target values/hyphens
                    # Apply number format only if it's a number
                    if isinstance(cell.value, (int, float)):
                        if cell.value == 0: # Display 0 as '-'
                           cell.value = '-'
                           cell.number_format = '@' # Text format for '-'
                        else:
                            cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
                    # No need for else, '-' is already centered


        # --- Adjust Column Widths (Reuse improved logic) ---
        for sheet_name in writer.sheets:
            sheet = writer.sheets[sheet_name]
            # Determine number column indices for this specific sheet
            current_number_cols_indices = []
            if sheet_name == summary_sheet_name:
                current_number_cols_indices = number_cols_indices
            elif sheet_name in daily_data_frames:
                 # Last column is the number column in daily tabs
                 current_number_cols_indices = [len(daily_data_frames[sheet_name].columns)]

            for col_idx_zero_based, column_cells in enumerate(sheet.columns):
                column_letter = get_column_letter(col_idx_zero_based + 1)
                max_length = 0
                is_number_col_flag = (col_idx_zero_based + 1) in current_number_cols_indices

                for cell in column_cells:
                    if sheet_name == summary_sheet_name and cell.row == 1: # Skip merged title row
                        header_cell_val = sheet.cell(row=3, column=col_idx_zero_based+1).value
                        max_length = max(max_length, len(str(header_cell_val or "")))
                        continue

                    try:
                        cell_str = str(cell.value or "")
                        cell_str_len = len(cell_str)

                        if is_number_col_flag and isinstance(cell.value, (int, float)) and cell.value != 0:
                           # Approx width with commas
                           num_commas = (len(str(int(cell.value))) - 1) // 3 if cell.value >= 1000 else 0
                           cell_str_len += num_commas
                           max_length = max(max_length, cell_str_len, 8) # Min width 8 for numbers
                        else:
                            max_length = max(max_length, cell_str_len)
                    except:
                        pass

                adjusted_width = max_length + 3
                sheet.column_dimensions[column_letter].width = adjusted_width

            # Specific overrides
            if sheet_name == summary_sheet_name:
                sheet.column_dimensions['B'].width = 25 # Agent Name
            # Make date columns slightly wider
            if sheet_name in daily_data_frames:
                 sheet.column_dimensions['A'].width = 12 # Date column

    print(f"Successfully created Excel file: {filename}")


# --- Run the Script ---
if __name__ == "__main__":
    input_month, input_year = get_month_year()
    generate_sales_target_excel(input_month, input_year)